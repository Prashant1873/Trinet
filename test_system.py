"""
TRINET (TM) Automated System Test Suite
Verifies all routes, HTML templates, static assets, APIs, and exports.
"""

import requests
import io
import openpyxl

BASE_URL = 'http://127.0.0.1:5000'

def run_tests():
    print("==================================================")
    print("TRINET (TM) Automated Test Suite")
    print("==================================================")

    # 1. HTML Index
    r = requests.get(BASE_URL + '/')
    assert r.status_code == 200, f"HTML index failed: {r.status_code}"
    assert "TRINET" in r.text
    print("[PASS] HTML SPA Shell loaded successfully (status 200)")

    # 2. Static CSS Files
    css_files = ['design-system.css', 'components.css', 'layout.css', 'map.css']
    for c in css_files:
        res = requests.get(f"{BASE_URL}/static/css/{c}")
        assert res.status_code == 200, f"Failed to load static/css/{c}"
    print(f"[PASS] All {len(css_files)} CSS files served correctly")

    # 3. Static JS Files
    js_files = ['app.js', 'map.js', 'filters.js', 'results.js', 'search.js', 'export.js', 'company.js', 'dashboard.js']
    for j in js_files:
        res = requests.get(f"{BASE_URL}/static/js/{j}")
        assert res.status_code == 200, f"Failed to load static/js/{j}"
    print(f"[PASS] All {len(js_files)} JS modules served correctly")

    # 4. Metadata
    r = requests.get(f"{BASE_URL}/api/metadata")
    assert r.status_code == 200
    meta = r.json()
    assert len(meta['industries']) >= 18
    assert len(meta['states']) >= 20
    print(f"[PASS] Metadata API returned {len(meta['industries'])} industries and {len(meta['states'])} states")

    # 5. Stats API
    r = requests.get(f"{BASE_URL}/api/stats")
    assert r.status_code == 200
    stats = r.json()
    assert stats['total_companies'] >= 500
    assert stats['total_facilities'] >= 800
    print(f"[PASS] Stats API: {stats['total_companies']} companies, {stats['total_facilities']} facilities mapped")

    # 6. Multi-facet Filter Query
    r = requests.get(f"{BASE_URL}/api/companies?industry=Automotive&scale=LARGE,ENTERPRISE&limit=5")
    assert r.status_code == 200
    comps = r.json()
    assert len(comps['data']) > 0
    print(f"[PASS] Multi-Filter API: {comps['pagination']['total']} Automotive Large/Enterprise companies found")

    # 7. Company Details API
    comp_id = comps['data'][0]['id']
    r = requests.get(f"{BASE_URL}/api/companies/{comp_id}")
    assert r.status_code == 200
    detail = r.json()
    assert detail['company']['company_name']
    assert detail['company'].get('email') is not None, "Company email missing"
    assert detail['company'].get('phone') is not None, "Company phone missing"
    assert len(detail['facilities']) > 0
    assert detail['facilities'][0].get('email') is not None, "Facility email missing"
    assert detail['facilities'][0].get('phone') is not None, "Facility phone missing"
    print(f"[PASS] Company Details API: {detail['company']['company_name']} ({len(detail['facilities'])} mapped sites) with Contact Intelligence ({detail['company']['email']}, {detail['company']['phone']})")

    # 8. Spatial Clustering API & GeoJSON
    r = requests.get(f"{BASE_URL}/api/facilities/geojson")
    assert r.status_code == 200
    geojson = r.json()
    assert geojson['type'] == 'FeatureCollection'
    assert len(geojson['features']) >= 1500
    print(f"[PASS] Native GeoJSON FeatureCollection API: {len(geojson['features'])} planar-locked features")

    # 9. AI Search NLP API
    r = requests.post(f"{BASE_URL}/api/ai/search", json={"query": "Pharma factories in Hyderabad with 500+ employees"})
    assert r.status_code == 200
    ai_res = r.json()
    assert ai_res.get('filters') is not None
    print(f"[PASS] AI NLP Search API: '{ai_res.get('explanation')}'")

    # 10. Excel Data Export (.xlsx) with Email & Phone
    r = requests.post(f"{BASE_URL}/api/export", json={"format": "xlsx"})
    assert r.status_code == 200
    assert len(r.content) > 10000
    # Validate with openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "Companies" in wb.sheetnames
    assert "Facilities" in wb.sheetnames
    
    ws_comp = wb["Companies"]
    comp_headers = [cell.value for cell in ws_comp[1]]
    assert "Email" in comp_headers, f"Email missing from Companies sheet headers: {comp_headers}"
    assert "Phone" in comp_headers, f"Phone missing from Companies sheet headers: {comp_headers}"
    
    ws_fac = wb["Facilities"]
    fac_headers = [cell.value for cell in ws_fac[1]]
    assert "Email" in fac_headers, f"Email missing from Facilities sheet headers: {fac_headers}"
    assert "Phone" in fac_headers, f"Phone missing from Facilities sheet headers: {fac_headers}"
    print(f"[PASS] Excel Export API: Generated {len(r.content):,} bytes (.xlsx) with 'Email' & 'Phone' columns verified in both sheets")

    # 11. CSV Data Export (.csv) with Email & Phone
    r = requests.post(f"{BASE_URL}/api/export", json={"format": "csv"})
    assert r.status_code == 200
    lines = r.text.splitlines()
    assert "Company Name" in lines[0]
    assert "Email" in lines[0]
    assert "Phone" in lines[0]
    print(f"[PASS] CSV Export API: Generated valid CSV stream with verified 'Email' & 'Phone' headers ({len(lines)} rows)")

    # 12. Discovery Coverage API
    r = requests.get(f"{BASE_URL}/api/discovery/coverage")
    assert r.status_code == 200
    cov = r.json()
    assert len(cov['coverage']) >= 20
    assert len(cov.get('corridors', [])) == 13
    print(f"[PASS] Discovery Coverage API: {len(cov['coverage'])} states tracked and {len(cov['corridors'])} industrial corridors mapped")

    # 13. Industrial Corridors Discovery & Scanning API
    r = requests.get(f"{BASE_URL}/api/discovery/corridors")
    assert r.status_code == 200
    corr_res = r.json()
    assert corr_res['total_corridors'] == 13
    print(f"[PASS] Industrial Corridors API: 13 National & Defence Industrial Corridors verified")

    r = requests.post(f"{BASE_URL}/api/discovery/corridor/scan", json={"corridor_code": "UPDIC"})
    assert r.status_code == 200
    scan_res = r.json()
    assert scan_res['corridor_code'] == 'UPDIC'
    assert len(scan_res['nodes_scanned']) > 0
    print(f"[PASS] Industrial Corridor Multi-Node Scan API: {scan_res['corridor_name']} scanned ({scan_res['total_new_companies']} new manufacturers, {scan_res['total_new_facilities']} facilities)")

    print("\n==================================================")
    print(">>> ALL 13 TRINET SYSTEM INTEGRATION TESTS PASSED! <<<")
    print("==================================================")

if __name__ == '__main__':
    run_tests()
