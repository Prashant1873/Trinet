"""
TRINET (TM) Data Exporter
Generates styled Excel (.xlsx) workbooks with Companies & Facilities sheets and CSV exports.
"""

import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_export(companies, facilities=None):
    """
    Generate an Excel workbook in memory with styled headers and data.
    """
    wb = openpyxl.Workbook()
    
    # ── Sheet 1: Companies ──
    ws_comp = wb.active
    ws_comp.title = "Companies"
    
    # Header styles
    header_fill = PatternFill(start_color="00A06C", end_color="00A06C", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    headers_comp = [
        "Company Name", "Industry", "Sub-Industry", "Email", "Phone", "City", "State", 
        "Established", "Website", "Scale", "Scale Score", "Employees", 
        "Facilities", "Exporter", "Public", "Verification", "Updated"
    ]
    
    ws_comp.append(headers_comp)
    for col_num, _ in enumerate(headers_comp, 1):
        cell = ws_comp.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_comp.row_dimensions[1].height = 26
        
    for r_idx, c in enumerate(companies, 2):
        row_data = [
            c.get('company_name', ''),
            c.get('industry', ''),
            c.get('sub_industry', ''),
            c.get('email', ''),
            c.get('phone', ''),
            c.get('headquarters_city', ''),
            c.get('headquarters_state', ''),
            c.get('establishment_year', ''),
            c.get('website', ''),
            c.get('company_scale', ''),
            c.get('scale_score', ''),
            c.get('employee_count', ''),
            c.get('facility_count', 1),
            "Yes" if c.get('is_exporter') else "No",
            "Yes" if c.get('is_public_company') else "No",
            c.get('verification_status', ''),
            c.get('updated_at', '')[:10] if c.get('updated_at') else ''
        ]
        ws_comp.append(row_data)
        for col_num in range(1, len(row_data) + 1):
            cell = ws_comp.cell(row=r_idx, column=col_num)
            cell.font = data_font
            cell.border = thin_border
            if col_num in (8, 11, 12, 13): # numeric columns
                cell.alignment = Alignment(horizontal="center")
                
    # Auto-adjust column widths
    for col in ws_comp.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_comp.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # ── Sheet 2: Facilities ──
    if facilities:
        ws_fac = wb.create_sheet(title="Facilities")
        headers_fac = [
            "Company Name", "Facility Name", "Facility Type", "Address", 
            "City", "State", "District", "PIN Code", "Email", "Phone", 
            "Latitude", "Longitude", "Google Rating", "Review Count", "Status"
        ]
        ws_fac.append(headers_fac)
        for col_num, _ in enumerate(headers_fac, 1):
            cell = ws_fac.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws_fac.row_dimensions[1].height = 26
            
        for r_idx, f in enumerate(facilities, 2):
            row_data = [
                f.get('company_name', ''),
                f.get('facility_name', ''),
                f.get('facility_type', ''),
                f.get('address', ''),
                f.get('city', ''),
                f.get('state', ''),
                f.get('district', ''),
                f.get('pincode', ''),
                f.get('email', ''),
                f.get('phone', ''),
                f.get('latitude', ''),
                f.get('longitude', ''),
                f.get('google_rating', ''),
                f.get('review_count', ''),
                f.get('operational_status', '')
            ]
            ws_fac.append(row_data)
            for col_num in range(1, len(row_data) + 1):
                cell = ws_fac.cell(row=r_idx, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                if col_num in (8, 10, 11, 12, 13, 14):
                    cell.alignment = Alignment(horizontal="center")
                    
        for col in ws_fac.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_fac.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_csv_export(companies):
    """
    Generate CSV string in memory for export.
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    headers = [
        "Company Name", "Industry", "Sub-Industry", "Email", "Phone", "City", "State", 
        "Year Established", "Website", "Scale", "Scale Score", "Employees", 
        "Facilities", "Exporter", "Public Company", "Verification Status", "Last Updated"
    ]
    writer.writerow(headers)
    
    for c in companies:
        writer.writerow([
            c.get('company_name', ''),
            c.get('industry', ''),
            c.get('sub_industry', ''),
            c.get('email', ''),
            c.get('phone', ''),
            c.get('headquarters_city', ''),
            c.get('headquarters_state', ''),
            c.get('establishment_year', ''),
            c.get('website', ''),
            c.get('company_scale', ''),
            c.get('scale_score', ''),
            c.get('employee_count', ''),
            c.get('facility_count', 1),
            "Yes" if c.get('is_exporter') else "No",
            "Yes" if c.get('is_public_company') else "No",
            c.get('verification_status', ''),
            c.get('updated_at', '')[:10] if c.get('updated_at') else ''
        ])
        
    output.seek(0)
    return output.getvalue().encode('utf-8')
